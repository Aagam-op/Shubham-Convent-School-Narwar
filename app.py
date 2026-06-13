import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook, Workbook
from flask import Flask, render_template_string, request, jsonify, send_from_directory

app = Flask(__name__)

# --- EXCEL SETUP ---
EXCEL_FILE = "inquiries.xlsx"

# Agar excel file pehle se nahi bani hai, toh nayi file banayein aur headers dalein
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    ws.append(["Name", "Phone", "Class"])  # Columns ke naam
    wb.save(EXCEL_FILE)

# --- 1. DATA DICTIONARY (Updated with Doctor Seminar & Poster) ---

# --- 1. DATA DICTIONARY (Updated with Doctor Seminar & Poster) ---
content_data = {
    'en': {
        'lang': 'en',
        'school_name': 'Shubham Convent School',
        'affiliation': 'Affiliated with MP Board',
        'nav_admission': 'Admission',
        'nav_legacy': 'Legacy',
        'nav_gallery': 'Gallery',
        'nav_find': 'Find Us',
        'sec_founder': 'Our Founder',
        'sec_gallery': 'Campus & Health Initiatives',
        'sec_location': 'Locate Us in Narwar',
        'logo_path': 'logo.png',
        'start_btn': 'Start Website',
        'stop_btn': 'Stop Presentation',
        'hero_img': 'School.png',
        'founder_img': 'founder1.png',
        'founder_name': 'Pandit Murarilal Jain',
        'founder_desc': 'Our founder, Pandit Murarilal Jain, has been leading 30 years of academic excellence in Narwar since 1994.',
        'gallery': [
            {'file': 'Poster.jpg', 'desc': 'Upcoming Programs & Events.'},
            {'file': 'Doctor_visit.jpg', 'desc': 'Every year, a seminar is conducted by a Doctor for children’s health.'},
            {'file': 'cmo_award.jpg', 'desc': 'Excellence Award from CMO.'},
            {'file': 'cultural_dance.jpg', 'desc': 'Cultural Performance & Heritage.'},
            {'file': 'group_photo.jpg', 'desc': 'Our Academic Toppers.'},
            {'file': 'cultural_event.jpg', 'desc': 'Celebrating National Festivals and Cultural Heritage.'}
        ],
        'admission_title': 'Admissions Open: Nursery to 8th',
        'inquiry_btn': 'Send Inquiry Now',
        'sending_btn': 'Processing...',
        'modal_name': 'Student Name', 'modal_phone': 'Phone Number', 'modal_grade': 'Select Class',
        'grades': ['Nursery', 'LKG', 'UKG', 'Class 1', 'Class 2', 'Class 3', 'Class 4', 'Class 5', 'Class 6', 'Class 7', 'Class 8'],
        'success_msg': 'Inquiry Sent Successfully!',
        'footer_id': 'School ID: 2026-Shubhamj646@gmail.com',
        'footer_contact': 'Contact: +91 8463001836',
        'footer_rights': 'All rights reserved @ 2026'
    },
    'hi': {
        'lang': 'hi',
        'school_name': 'शुभम कॉन्वेंट स्कूल',
        'affiliation': 'मध्य प्रदेश बोर्ड (MP Board) से संबंध',
        'nav_admission': 'प्रवेश',
        'nav_legacy': 'विरासत',
        'nav_gallery': 'गैलरी',
        'nav_find': 'पता',
        'sec_founder': 'हमारे संस्थापक',
        'sec_gallery': 'कैंपस और स्वास्थ्य पहल',
        'sec_location': 'नरवर में हमारा स्थान',
        'logo_path': 'logo.png',
        'start_btn': 'वेबसाइट शुरू करें',
        'stop_btn': 'प्रस्तुति रोकें',
        'hero_img': 'School.png',
        'founder_img': 'founder1.png',
        'founder_name': 'पंडित मुरारीलाल जैन',
        'founder_desc': 'हमारे संस्थापक, पंडित मुरारीलाल जैन, 1994 से नरवर में 30 वर्षों की शैक्षणिक उत्कृष्टता का नेतृत्व कर रहे हैं।',
        'gallery': [
            {'file': 'Poster.jpg', 'desc': 'आगामी कार्यक्रम और पोस्टर।'},
            {'file': 'Doctor_visit.jpg', 'desc': 'हर साल बच्चों के स्वास्थ्य के लिए डॉक्टर द्वारा सेमिनार आयोजित किया जाता है।'},
            {'file': 'cultural_event.jpg', 'desc': 'राष्ट्रीय त्योहारों और सांस्कृतिक विरासत का जश्न।'},
            {'file': 'cmo_award.jpg', 'desc': 'CMO से उत्कृष्टता पुरस्कार।'},
            {'file': 'cultural_dance.jpg', 'desc': 'सांस्कृतिक प्रदर्शन और विरासत।'},
            {'file': 'group_photo.jpg', 'desc': 'हमारे शैक्षणिक टॉपर्स।'}
        ],
        'admission_title': 'प्रवेश प्रारंभ: नर्सरी से 8वीं',
        'inquiry_btn': 'पूछताछ भेजें',
        'sending_btn': 'भेज रहे हैं...',
        'modal_name': 'छात्र का नाम', 'modal_phone': 'फोन नंबर', 'modal_grade': 'कक्षा चुनें',
        'grades': ['नर्सरी', 'एल.के.जी.', 'यू.के.जी.', 'कक्षा 1', 'कक्षा 2', 'कक्षा 3', 'कक्षा 4', 'कक्षा 5', 'कक्षा 6', 'कक्षा 7', 'कक्षा 8'],
        'success_msg': 'पूछताछ सफलतापूर्वक भेजी गई!',
        'footer_id': 'स्कूल आईडी: 2026-शुभम',
        'footer_contact': 'संपर्क: +91 8463001836',
        'footer_rights': 'सर्वाधिकार सुरक्षित @ 2026'
    }
}

# --- 2. HTML TEMPLATE (With AOS Effects Enabled) ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="{{ content.lang }}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>{{ content.school_name }}</title>
    
    <link href="https://unpkg.com/aos@2.3.1/dist/aos.css" rel="stylesheet">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/gsap/3.12.2/gsap.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/gsap/3.12.2/ScrollToPlugin.min.js"></script>
    
    <style>
        :root { --red: #8B0000; --gold: #D4AF37; --cream: #fffaf5; }
        * { box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
        body { 
    font-family: 'Segoe UI', sans-serif; 
    margin: 0; 
    background: var(--cream); 
    padding-top: 70px; 
    overflow-x: hidden; 
    width: 100vw; 
}
        
        .navbar { position: fixed; top: 0; width: 100%; height: 70px; background: white; display: flex; justify-content: space-between; align-items: center; padding: 0 5%; z-index: 1000; border-bottom: 3px solid var(--gold); }
        .nav-links a { text-decoration: none; color: var(--red); font-weight: bold; margin-left: 15px; }

        .hero { height: 80vh; background: linear-gradient(rgba(0,0,0,0.6), rgba(0,0,0,0.6)), url('/static/images/{{ content.hero_img }}'); background-size: cover; background-position: center; display: flex; align-items: center; justify-content: center; color: white; text-align: center; }
        
        .section { padding: 60px 8%; text-align: center; }
        .founder-card { display: flex; align-items: center; justify-content: center; gap: 40px; flex-wrap: wrap; background: white; padding: 30px; border-radius: 15px; box-shadow: 0 5px 15px rgba(0,0,0,0.05); border-left: 5px solid var(--red); }
        .founder-card img { width: 220px; height: 220px; border-radius: 50%; border: 5px solid var(--gold); object-fit: cover; object-position: top center; }

    .slider-container {
            position: relative;
            max-width: 900px;
            margin: 0 auto;
            overflow: hidden;
            border-radius: 20px;
            box-shadow: 0 15px 35px rgba(0,0,0,0.12);
            background: #fff;
            padding: 15px;
        }
        .slider-wrapper {
            display: flex;
            transition: transform 0.8s cubic-bezier(0.25, 1, 0.5, 1);
            width: 100%;
        }
        .slider-item {
            min-width: 100%;
            box-sizing: border-box;
            padding: 0 10px;
            text-align: center;
            position: relative;
        }
        .slider-item img {
            width: 100%;
            height: 480px;
            object-fit: cover;
            border-radius: 15px;
            transition: transform 0.5s ease;
        }
        .slider-container:hover .slider-item img {
            transform: scale(1.02);
        }
        .slider-desc {
            font-weight: 700;
            color: var(--red);
            margin-top: 20px;
            font-size: 1.2rem;
            letter-spacing: 0.5px;
        }
        /* PREMIUM DOTS STYLING */
        .slider-dots {
            display: flex;
            justify-content: center;
            gap: 10px;
            margin-top: 15px;
        }
        .dot {
            width: 12px;
            height: 12px;
            border-radius: 50%;
            background: #ddd;
            cursor: pointer;
            transition: all 0.4s ease;
        }
        .dot.active {
            background: var(--red);
            width: 28px;
            border-radius: 10px;
        }
        @media (max-width: 768px) {
            .slider-container { padding: 10px; }
            .slider-item img { height: 300px; }
            .slider-desc { font-size: 1rem; margin-top: 12px; }
            .dot { width: 9px; height: 9px; }
            .dot.active { width: 20px; }
        }
        @media (max-width: 768px) {
            .slider-container { padding: 10px; }
            .slider-item img { height: 300px; }
            .slider-desc { font-size: 1rem; margin-top: 12px; }
            .dot { width: 9px; height: 9px; }
            .dot.active { width: 20px; }
        }
        .map-container { width: 100%; height: 450px; margin-top: 30px; border-radius: 15px; overflow: hidden; border: 3px solid var(--gold); box-shadow: 0 10px 20px rgba(0,0,0,0.1); }

        .admission-box { background: white; padding: 40px; border-radius: 20px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); max-width: 500px; margin: 0 auto; border-top: 10px solid var(--red); }
        input, select { width: 100%; padding: 14px; margin: 10px 0; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; }
        .submit-btn { background: var(--red); color: white; border: none; padding: 15px; border-radius: 8px; width: 100%; font-weight: bold; cursor: pointer; transition: 0.3s; }
        .submit-btn:hover { background: #660000; }

        footer { background: #000; color: white; padding: 50px 8%; text-align: center; margin-top: 50px; }
        .footer-highlight { color: var(--gold); font-weight: bold; font-size: 1.2rem; display: block; margin-bottom: 10px; text-transform: uppercase; }

        .start-btn { position: fixed; bottom: 25px; left: 25px; z-index: 100; background: var(--red); color: white; padding: 12px 25px; border-radius: 50px; border: 2px solid var(--gold); font-weight: bold; cursor: pointer; box-shadow: 0 5px 15px rgba(0,0,0,0.3); }
        .whatsapp { position: fixed; bottom: 25px; right: 25px; z-index: 100; }
       /* CALL BUTTON STYLING */
.call-btn {
    position: fixed;
    bottom: 100px;
    right: 25px;
    z-index: 1000;
    background: transparent; /* This is the fix - no extra color here */
    border: none;
    transition: transform 0.3s ease;
}

.call-btn img {
    width: 60px; /* This makes it match the WhatsApp size */
    height: auto;
    /* Optional: adds a shadow to the icon itself since the background is gone */
    filter: drop-shadow(0 4px 6px rgba(0,0,0,0.2)); 
}

/* WHATSAPP POSITION CHECK */
.whatsapp {
    position: fixed;
    bottom: 25px;
    right: 25px;
    z-index: 1000;
}

/* TOP BUTTON STYLING */
.top-btn {
    position: fixed;
    bottom: 175px;
    right: 25px;
    z-index: 1000;
    background: var(--gold);
    color: black;
    border: none;
    padding: 10px 15px;
    border-radius: 5px;
    cursor: pointer;
    display: none;
    font-weight: bold;
    box-shadow: 0 4px 6px rgba(0,0,0,0.2);
}
    </style>
</head>
<body>

    <nav class="navbar">
        <div style="display: flex; align-items: center;">
            <img src="/static/images/{{ content.logo_path }}" height="50">
            <div style="margin-left: 12px;">
                <span style="color: var(--red); font-weight: 900; display: block; font-size: 1.1rem;">{{ content.school_name }}</span>
                <small style="color: #555; font-weight: bold;">{{ content.affiliation }}</small>
            </div>
        </div>
       <div class="nav-links">
            <a href="#founder">{{ content.nav_legacy }}</a>
            <a href="#gallery">{{ content.nav_gallery }}</a>
            <a href="#location">{{ content.nav_find }}</a>
            <a href="#admission" style="background: var(--red); color: white; padding: 10px 20px; border-radius: 5px;">{{ content.nav_admission }}</a>
            <a href="?lang=hi" style="margin-left: 20px;">हिन्दी</a>
            <a href="?lang=en">EN</a>
        </div>
    </nav>

    <button class="start-btn" id="start-web">▶ {{ content.start_btn }}</button>

    <a href="https://wa.me/918463001836" class="whatsapp" target="_blank">
        <img src="https://upload.wikimedia.org/wikipedia/commons/6/6b/WhatsApp.svg" width="65">
    </a>

    <header class="hero">
        <div data-aos="zoom-in" data-aos-duration="1500">
            <h1 style="font-size: 3.5rem; margin: 0;">{{ content.school_name }}</h1>
            <p style="letter-spacing: 4px; color: var(--gold); font-weight: bold; font-size: 1.2rem;">NARWAR, SHIVPURI | ESTD 1994</p>
        </div>
    </header>
    <section class="section" id="founder">
    <div class="founder-card" data-aos="fade-right">
            <img src="/static/images/{{ content.founder_img }}" alt="Founder">
            <div style="text-align: left; max-width: 550px;">
                <h2 style="color: var(--red);">{{ content.sec_founder }}</h2>
<h2 style="color: var(--red); font-size: 2rem;">{{ content.founder_name }}</h2>
                <p style="font-size: 1.1rem; line-height: 1.6;">{{ content.founder_desc }}</p>
            </div>
        </div>
    </section>

    <section class="section" id="gallery" style="background: #f9f9f9;">
    <h2 style="color: var(--red); margin-bottom: 40px;" data-aos="fade-up">{{ content.sec_gallery }}</h2>
    <div class="slider-container" data-aos="fade-up">
        <div class="slider-wrapper" id="sliderTrack">
            {% for item in content.gallery %}
            <div class="slider-item">
                <img src="/static/images/{{ item.file }}" alt="Gallery Image">
                <p class="slider-desc">{{ item.desc }}</p>
            </div>
            {% endfor %}
        </div>
        <div class="slider-dots">
            {% for item in content.gallery %}
            <span class="dot {% if loop.first %}active{% endif %}" onclick="currentSlide({{ loop.index0 }})"></span>
            {% endfor %}
        </div>
    </div>
</section>

    <section class="section" id="location">
    <h2 style="color: var(--red);" data-aos="fade-down">{{ content.sec_location }}</h2>
        <p style="font-weight: bold; color: #555;">Landmark: Near Narwar Fort Road & Gas Agency</p>
        <div class="map-container" data-aos="zoom-in">
            <iframe src="https://www.google.com/maps/embed?pb=!1m18!1m12!1m3!1d14343.376823351!2d77.9048384!3d25.6515865!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!3m3!1m2!1s0x39773950f280a96d%3A0x60f06871a33719b0!2sNarwar%2C%20Madhya%20Pradesh%20473880!5e0!3m2!1sen!2sin!4v1715690000000!5m2!1sen!2sin" width="100%" height="100%" style="border:0;" allowfullscreen="" loading="lazy" scrolling="no" referrerpolicy="no-referrer-when-downgrade"></iframe>
        </div>
    </section>

    <section class="section" id="admission">
        <div class="admission-box" data-aos="flip-up">
            <h2 style="color: var(--red);">{{ content.admission_title }}</h2>
            <form id="inquiry-form">
                <input type="text" id="student_name" name="name" placeholder="{{ content.modal_name }}" required>
                <input type="tel" id="student_phone" name="phone" placeholder="{{ content.modal_phone }}" pattern="[0-9]{10}" required>
                <select id="student_grade" name="grade" required>
                    <option value="" disabled selected>{{ content.modal_grade }}</option>
                    {% for grade in content.grades %}
                    <option value="{{ grade }}">{{ grade }}</option>
                    {% endfor %}
                </select>
                <button type="submit" id="submit-btn" class="submit-btn">{{ content.inquiry_btn }}</button>
            </form>
        </div>
    </section>

    <footer>
        <div data-aos="fade-up">
            <span class="footer-highlight">{{ content.footer_id }}</span>
            <span class="footer-highlight">{{ content.footer_contact }}</span>
            <p style="margin: 20px 0; font-weight: bold; color: var(--gold);">{{ content.affiliation }}</p>
            <hr style="border: 0; border-top: 1px solid #333; margin: 25px 0;">
            <p style="font-size: 0.9rem; opacity: 0.7;">{{ content.footer_rights }}</p>
        </div>
    </footer>
    <a href="tel:+918463001836" class="call-btn">
      <img src="https://cdn-icons-png.flaticon.com/512/3616/3616215.png" width="40">
    </a>
    <button id="topBtn" class="top-btn">↑</button>
    

    <script src="https://unpkg.com/aos@2.3.1/dist/aos.js"></script>
    <script>
        // Initialize AOS with proper settings to ensure effects are visible
        AOS.init({
            duration: 1000,
            once: false, 
            mirror: true
        });

        // --- GLITCH PREVENTION FUNCTIONS ---
        const stopScroll = () => {
            if (isScrolling) {
                if(scrollTween) scrollTween.kill();
                toggle(false);
            }
        };

        // Scroll/Touch/Keyboard events par auto-scroll rokne ke liye
        window.addEventListener('wheel', stopScroll, { passive: true });
        window.addEventListener('touchstart', stopScroll, { passive: true });
        window.addEventListener('keydown', stopScroll, { passive: true });

        // --- STYLISH GALLERY SLIDER WITH ACTIVE DOT SYNC ---

       // --- STYLISH GALLERY SLIDER WITH ACTIVE DOT SYNC ---
    const track = document.getElementById('sliderTrack');
    const items = document.querySelectorAll('.slider-item');
    const dots = document.querySelectorAll('.dot');
    let currentIndex = 0;
    const totalItems = items.length;

    function updateSlider(index) {
        currentIndex = index;
        if(track) {
            track.style.transform = `translateX(-${currentIndex * 100}%)`;
        }
        // Update indicator states seamlessly
        dots.forEach(dot => dot.classList.remove('active'));
        if(dots[currentIndex]) {
            dots[currentIndex].classList.add('active');
        }
    }

    function slideNext() {
        let nextIndex = (currentIndex + 1) % totalItems;
        updateSlider(nextIndex);
    }

    // Allow clicking directly on a dot to navigate
    function currentSlide(index) {
        updateSlider(index);
    }

    setInterval(slideNext, 4000);

        let isScrolling = false;
        let scrollTween;

        // Smooth nav links scroll
        document.querySelectorAll('.nav-links a').forEach(anchor => {
            anchor.addEventListener('click', function(e) {
                if(this.getAttribute('href').startsWith('#')) {
                    e.preventDefault();
                    const target = document.querySelector(this.getAttribute('href'));
                    gsap.to(window, { duration: 1, scrollTo: target, ease: "power2.inOut" });
                }
            });
        });

        document.getElementById('start-web').addEventListener('click', function() {
            if (!isScrolling) {
                // Auto-scroll set to 25 seconds
                scrollTween = gsap.to(window, { 
                    duration: 25, 
                    scrollTo: document.body.scrollHeight, 
                    ease: "none", 
                    overwrite: true,
                   
                    onComplete: () => { toggle(false); }
                });
                toggle(true);
            } else {
                scrollTween.kill();
                toggle(false);
            }
        });

        function toggle(s) {
            isScrolling = s;
            document.getElementById('start-web').innerText = s ? "⏹ {{ content.stop_btn }}" : "▶ {{ content.start_btn }}";
        }

        // Go to Top Logic
        const topBtn = document.getElementById("topBtn");
        window.onscroll = () => {
            if (document.body.scrollTop > 300 || document.documentElement.scrollTop > 300) {
                topBtn.style.display = "block";
            } else {
                topBtn.style.display = "none";
            }
        };
        topBtn.addEventListener("click", () => {
            gsap.to(window, { duration: 1, scrollTo: 0, ease: "power2.inOut" });
        });

        document.getElementById('inquiry-form').addEventListener('submit', async (e) => {
            e.preventDefault();
            const btn = document.getElementById('submit-btn');
            btn.disabled = true;
            btn.innerText = "{{ content.sending_btn }}";
            
            // Explicitly selecting data to prevent any blank transfers
            const formData = new FormData();
      formData.append('name', document.getElementById('student_name').value);
formData.append('phone', document.getElementById('student_phone').value);
formData.append('grade', document.getElementById('student_grade').value);
            try {
                const response = await fetch('/submit', { method: 'POST', body: formData });
                if(response.ok) {
                    alert("{{ content.success_msg }}");
                    e.target.reset();
                } else {
                    alert("Server Error: Could not save data.");
                }
            } catch (err) {
                alert("Network Error: Connection failed.");
            }
            
            btn.disabled = false;
            btn.innerText = "{{ content.inquiry_btn }}";
        });
    </script>
</body>
</html>
"""
# --- 3. BACKEND (Flask) ---
@app.route('/')
def index():
    lang = request.args.get('lang', 'en')
    return render_template_string(HTML_TEMPLATE, content=content_data.get(lang, content_data['en']))
@app.route('/submit', methods=['POST'])
def submit():
    name = request.form.get('name')
    phone = request.form.get('phone')
    grade = request.form.get('grade')
    
    print(f"--- SUBMIT TRIGGERED ---")

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, phone, grade])
        wb.save(EXCEL_FILE)
        
        # Yeh line aapko batayegi ki asal mein file kahaan save hui hai
        import os
        print(f"🎯 FILE SAVED AT: {os.path.abspath(EXCEL_FILE)}")
        
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"❌ EXCEL SAVE ERROR: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/static/images/<path:filename>')
def custom_static(filename):
    return send_from_directory('static/images', filename)

if __name__ == '__main__':
    if not os.path.exists('static/images'):
        os.makedirs('static/images')
    app.run(debug=True, port=5001)